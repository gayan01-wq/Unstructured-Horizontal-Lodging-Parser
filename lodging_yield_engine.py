import pandas as pd
import numpy as np
import os
from google.colab import files
from IPython.display import HTML, display
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference

def run_wyndham_ultimate_compset_comparison_engine():
    HOTEL_NAME = "Wyndham Garden Salalah Mirbat"
    MY_INVENTORY = 237  
    
    COMPSET_CAPACITIES = {
        "Millennium Resort Salalah": 285,
        "Salalah Rotana Resort": 422,
        "Intercity Hotel Salalah": 70,
        "Salalah Gardens Hotel": 168,
        "Crowne Plaza Resort Salalah": 153,
        "Alila Hinu Bay Salalah": 112
    }
    TOTAL_MARKET_KEYS = sum(COMPSET_CAPACITIES.values()) + MY_INVENTORY
    
    revenue_inputs = {
        "Jun 2026": {"days": 30, "sold": 176, "rev": 2555, "market_occ": 25.0},
        "Jul 2026": {"days": 31, "sold": 680, "rev": 26876, "market_occ": 18.0},
        "Aug 2026": {"days": 31, "sold": 1950, "rev": 74100, "market_occ": 32.0}
    }
    
    processed_rows = []
    
    for period, data in revenue_inputs.items():
        days = data["days"]
        my_capacity = MY_INVENTORY * days
        my_occ = (data["sold"] / my_capacity) * 100
        adr = data["rev"] / data["sold"] if data["sold"] > 0 else 0.0
        
        market_capacity_pool = TOTAL_MARKET_KEYS * days
        fair_share_target = (market_capacity_pool * (data["market_occ"] / 100)) * (my_capacity / market_capacity_pool)
        variance = data["sold"] - fair_share_target
        
        if adr > 39.0 and my_occ < 15.0:
            yield_friction = "HIGH FRICTION (Rate is choking booking velocity; market resistance met)"
        elif adr > 35.0 and my_occ >= 30.0:
            yield_friction = "YIELD SWEET-SPOT (Premium pricing accepted by market; maximum capture)"
        else:
            yield_friction = "LOW FRICTION (Soft rate structure; rooms are filling with zero market resistance)"
        
        processed_rows.append({
            "Period": period,
            "Wyndham Occ %": round(my_occ, 1),
            "Market Occ %": round(data["market_occ"], 1),
            "Fair Share Target (Rooms)": round(fair_share_target, 0),
            "Actual Captured (Rooms)": data["sold"],
            "Market Share Variance": round(variance, 0),
            "ADR (OMR)": round(adr, 2),
            "Yield Friction Index": yield_friction
        })
        
    df_output = pd.DataFrame(processed_rows)
    output_filename = "Wyndham_Final_Compset_Comparison_Report.xlsx"
    
    if os.path.exists(output_filename):
        os.remove(output_filename)
        
    # =========================================================================
    # 📂 BUILD THE EXCEL WORKBOOK WITH SIDE-BY-SIDE COMPSET BARS
    # =========================================================================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Performance"
    ws.views.sheetView[0].showGridLines = True
    
    headers = list(df_output.columns)
    ws.append(headers)
    for index, row in df_output.iterrows():
        ws.append(list(row))
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # 1. Primary Bar Chart: Compare Fair Share vs Actual Captured side-by-side
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10  
    chart1.title = "Wyndham Actual Bookings vs Market Fair Share Baseline"
    chart1.y_axis.title = "Room Nights"
    chart1.x_axis.title = "Period"
    
    # Select Columns D & E (Fair Share Target & Actual Captured)
    volume_data = Reference(ws, min_col=4, min_row=1, max_col=5, max_row=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart1.add_data(volume_data, titles_from_data=True)
    chart1.set_categories(cats)
    
    # 2. Secondary Line Chart: ADR Price Curve
    chart2 = LineChart()
    adr_data = Reference(ws, min_col=7, min_row=1, max_row=4)
    chart2.add_data(adr_data, titles_from_data=True)
    
    chart2.y_axis.axId = 200
    chart2.y_axis.title = "ADR (OMR)"
    chart2.y_axis.crosses = "max"
    
    # Color the line chart Crimson Red
    s2 = chart2.series[0]
    s2.graphicalProperties.line.solidFill = "D93025"
    s2.graphicalProperties.line.width = 35000         
    
    # Combine charts
    chart1 += chart2
    
    chart1.height = 14
    chart1.width = 20
    ws.add_chart(chart1, "A8")  
    
    # Compset Setup Tab
    ws2 = wb.create_sheet(title="Calibrated Market Supply")
    ws2.append(["Compset Hotel Name", "Physical Rooms Base"])
    for hotel, inventory in COMPSET_CAPACITIES.items():
        ws2.append([hotel, inventory])
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 22
    
    wb.save(output_filename)
    
    print("\n" + "="*85)
    print(f"🎉 EXCEL GENERATED WITH SIDE-BY-SIDE COMPSET VISUAL COMPARISON!")
    print("="*85)
    
    link_html = f'<a href="/files/{output_filename}" download="{output_filename}" style="font-size:15px; font-weight:bold; color:#ffffff; background-color:#1a73e8; padding:12px 24px; border-radius:4px; text-decoration:none; display:inline-block; margin-top:10px; box-shadow: 0 2px 4px rgba(0,0,0,0.15);">➔ Download Final Presentation Report</a>'
    display(HTML(link_html))
    
    try: files.download(output_filename)
    except: pass

run_wyndham_ultimate_compset_comparison_engine()
